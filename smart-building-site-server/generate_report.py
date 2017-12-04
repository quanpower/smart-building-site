# -*- coding: utf-8 -*- 

from openpyxl import load_workbook
from openpyxl import Workbook

import datetime


def generate_datetime_list(startTime='2017-10-20 00:12:13', endTime='2017-10-22 23:12:13'):

	startTime = datetime.datetime.strptime(startTime, "%Y-%m-%d %H:%M:%S")
	endTime = datetime.datetime.strptime(endTime, "%Y-%m-%d %H:%M:%S")

	startDateTime = startTime
	datetime_list = []
	while startDateTime < endTime:
		datetime_list.append(startDateTime)
		startDateTime = startDateTime + datetime.timedelta(hours=1)

	datetime_str_list = []
	for i in range(len(datetime_list)):
		datetime_str_list.append(datetime_list[i].strftime("%Y-%m-%d %H:%M:%S"))

	print('---------datetime_str_list-----------')
	print(datetime_str_list)

	return datetime_list

def generate_report(nodeAddr='112',startTime='2017-10-20 00:12:13',endTime='2017-10-22 23:12:13'):
	wb = load_workbook('concrete_template.xlsx')  #加载一个工作簿
	print(wb.get_sheet_names())

	# wb = Workbook()
	#创建一个工作簿
	ws = wb.active
	#至少建立一个工作表
	#设置表的名字
	ws.sheet_properties.tabColor = "2072BA"
	#改变表选项卡的颜色
	ws["I8"] = 38.1
	print(ws["I8"].value)

	ws.cell(row=8, column=9, value=29.1)
	ws.cell(row=9, column=9, value=48.9)
	ws.cell(row=10, column=9, value=40.3)


	print(ws["I8"].value)

	wb.save(nodeAddr + '_' + startTime + '_' + endTime + '_' "sample.xlsx")

def main():
	generate_report()

if __name__ == '__main__':
	main()