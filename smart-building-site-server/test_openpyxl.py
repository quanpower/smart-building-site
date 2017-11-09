# -*- coding: utf-8 -*- 

from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook('concrete_template.xlsx')  #加载一个工作簿
print(wb.get_sheet_names())

# wb = Workbook()
#创建一个工作簿
ws = wb.active
#至少建立一个工作表
#设置表的名字
ws.sheet_properties.tabColor = "2072BA"
#改变表选项卡的颜色
ws["I8"] = 38.1
print(ws["I8"].value)

ws.cell(row=8, column=9, value=29.1)
ws.cell(row=9, column=9, value=48.9)
ws.cell(row=10, column=9, value=40.3)


print(ws["I8"].value)

wb.save("sample1-1.xlsx")